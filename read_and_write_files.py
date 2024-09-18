# TASK-2: CREATE 3 TYPES OF FILES:
'''1.JSON FILE
   2.CSV FILE
   3.EXCEL FILE   
   (Read and write your data-name, email, college name in all three files)
'''

# JSON
import json

dic= {
    "name":"Jyoti",
    "college_name":"ITS Ghaziabad",
    "email":"jyotipanchal2808@gmail.com"
}

# Write JSON
with open("user.json", "w") as json_file:
    json.dump(dic, json_file)

# Read JSON
with open("user.json", "r") as json_file:
    print("JSON Data:", json_file.read())


# CSV
import csv

header = ['name','college_name','email']
data = [
    ['Jyoti','ITS Ghaziabad','jyotipanchal2808@gmail.com']
]

# Write CSV
with open ('user.csv','w',newline="") as file:
    csvwriter=csv.writer(file)
    csvwriter.writerow(header)
    csvwriter.writerows(data)
   
# Read CSV
with open('user.csv','r') as csvfile:
    csv_reader=csv.reader(csvfile)
    print("CSV Data:")
    for row in csv_reader:
        print(row)

# EXCEL
import openpyxl

# Write EXCEL
wb=openpyxl.Workbook()
ws=wb.active
ws["A1"]="NAME"
ws["B1"]="COLLEGE NAME"
ws["C1"]="EMAIL"
ws["A2"]="Jyoti"
ws["B2"]="ITS Ghaziabad"
ws["C2"]="jyotipanchal2808@gmail.com"
wb.save("user.xlsx")

#Read EXCEL
wb=openpyxl.load_workbook("user.xlsx")
ws=wb.active
print("Excel Data:")
for row in ws.iter_rows(values_only=True):
    print(row)